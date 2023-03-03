from django.shortcuts import render

import requests
import openpyxl

import os
from django.http import FileResponse

import mimetypes


def get_publication_info(url):
    response = requests.get(url)
    publications_info = []
    if response.status_code == 200:
        data = response.json()
        entries = data.get("search-results").get("entry")
        for entry in entries:
            title = entry.get("dc:title")
            creator = entry.get("dc:creator")
            publicationName = entry.get("prism:publicationName")
            try:
                affil_name = entry["affiliation"][0]["affilname"]
                affiliation_city = entry["affiliation"][0]["affiliation-city"]
                affiliation_country = entry["affiliation"][0]["affiliation-country"]
            except Exception as e:
                pass
            articles_info = {
                "title": title,
                "creator": creator,
                "publicationName": publicationName,
                "affil_name": affil_name,
                "affiliation_city": affiliation_city,
                "affiliation_country": affiliation_country
            }

            publications_info.append(articles_info)
    else:
        print("Request failed with status code", response.status_code)

    return publications_info


def create_exl(lastname, firstname):
    author_search_url = f"https://api.elsevier.com/content/search/author?query=AUTHLASTNAME(%22{lastname}%22)%20AND%20AUTHFIRST(%22{firstname}%22)&apiKey=f0097b56b4ed057a6d01f4a20620c3d0"

    response = requests.get(author_search_url)
    if response.status_code == 200:
        data = response.json()
        id = data.get("search-results").get("entry")[0].get("dc:identifier")
        id = id.replace('AUTHOR_ID:', '')

    # AUID = int(input())
    # url = f"https://api.elsevier.com/content/search/scopus?query=AU-ID({AUID})&apiKey=f0097b56b4ed057a6d01f4a20620c3d0"
    url = f"https://api.elsevier.com/content/search/scopus?query=AU-ID({id})&apiKey=f0097b56b4ed057a6d01f4a20620c3d0"
    # url = f"https://api.elsevier.com/content/search/scopus?query=AU-ID(57209645760)&apiKey=f0097b56b4ed057a6d01f4a20620c3d0"
    publication_info = get_publication_info(url)

    for i in range(len(publication_info)):
        print(publication_info[i].get("title"))
        print(publication_info[i].get("creator"))
        print(publication_info[i].get("publicationName"))
        print(publication_info[i].get("affil_name"))
        print(publication_info[i].get("affiliation_city"))
        print(publication_info[i].get("affiliation_country"))
        print("--------------------------")

    # Open an existing Excel file or create a new one
    wb = openpyxl.Workbook()

    # Create two new sheets
    sheet1 = wb.create_sheet("Sheet1")
    sheet2 = wb.create_sheet("Sheet2")

    # Select the active sheet (Sheet1)
    sheet = wb.active

    # Write headers to the active sheet
    sheet['A1'] = 'Название статьи'
    for i in range(len(publication_info)):
        sheet[f'A{i + 2}'] = publication_info[i].get("title")
    sheet['B1'] = 'Автор'
    for i in range(len(publication_info)):
        sheet[f'B{i + 2}'] = publication_info[i].get("creator")
    sheet['C1'] = 'Журнал'
    for i in range(len(publication_info)):
        sheet[f'C{i + 2}'] = publication_info[i].get("publicationName")
    sheet['D1'] = 'Университет'
    for i in range(len(publication_info)):
        sheet[f'D{i + 2}'] = publication_info[i].get("affil_name")
    sheet['E1'] = 'Город'
    for i in range(len(publication_info)):
        sheet[f'E{i + 2}'] = publication_info[i].get("affiliation_city")
    sheet['F1'] = 'Страна'
    for i in range(len(publication_info)):
        sheet[f'F{i + 2}'] = publication_info[i].get("affiliation_country")
    sheet['G1'] = 'Оттиск прикрепить'
    sheet['H1'] = 'Количество цитирование Скопус/ веб оф сайнс'
    sheet['I1'] = 'Тип автора'
    sheet['J1'] = 'Количество авторов'
    sheet['D2'] = 'Поиск на сайте или или Название сайте ДА/нет'
    sheet['E2'] = 'Поиск на сайте или или Название сайте ДА/нет'
    sheet['F2'] = 'Поиск на сайте или или Название сайте ДА/нет'
    sheet['I2'] = 'Первый автор Автор кореспондент Соавтор (низбадающий список)'

    sheet = sheet1
    sheet['A1'] = 'Название Свидетельства'
    sheet['B1'] = 'Данные о свидетельстве'
    sheet['C1'] = 'авторы'
    sheet['D1'] = 'Прикрепить копию пдф'

    sheet = sheet2
    sheet['A1'] = 'Название '
    sheet['B1'] = 'Год выпуска, издательство.'
    sheet['C1'] = 'Количество страниц'
    sheet['C1'] = 'Прикрепить оттиск пдф'
    sheet['A2'] = 'Чтение с вложеноого файла автоматический'
    sheet['B2'] = 'Чтение с вложеноого файла автоматический'

    # Save the changes to the Excel file
    wb.save('media/data.xlsx')
    return publication_info


def search(request):
    if request.method == 'GET':
        return render(request, 'search.html', {})

    elif request.method == 'POST':
        lastname = request.POST['query']
        firstname = request.POST['another_query']

        info = create_exl(lastname, firstname)

        return render(request, 'search_results.html', {'query': info})


def download_xls(request):
    # Set the path to the XLSX file on disk
    filepath = 'media/data.xlsx'

    # Determine the MIME type based on the file extension
    mimetype, _ = mimetypes.guess_type(filepath)

    # Set the Content-Type header to the MIME type
    response = FileResponse(open(filepath, 'rb'), content_type=mimetype)

    # Set the Content-Disposition header to force download
    response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(filepath)

    return response
