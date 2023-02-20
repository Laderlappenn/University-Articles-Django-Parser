import openpyxl
from pars import get_publication_info

# AUID = int(input())
# url = f"https://api.elsevier.com/content/search/scopus?query=AU-ID({AUID})&apiKey=f0097b56b4ed057a6d01f4a20620c3d0"
url = "https://api.elsevier.com/content/search/scopus?query=AU-ID(57212065112)&apiKey=f0097b56b4ed057a6d01f4a20620c3d0"
publication_info = get_publication_info(url)


for i in range(len(publication_info)):
    print(publication_info[i].get("title"))
    print(publication_info[i].get("creator"))
    print(publication_info[i].get("publicationName"))
    print(publication_info[i].get("affil_name"))
    print(publication_info[i].get("affiliation_city"))
    print(publication_info[i].get("affiliation_country"))
    print("--------------------------")


# Open an existing Excel file or create a new one
wb = openpyxl.Workbook()

# Create two new sheets
sheet1 = wb.create_sheet("Sheet1")
sheet2 = wb.create_sheet("Sheet2")

# Select the active sheet (Sheet1)
sheet = wb.active

# Write headers to the active sheet
sheet['A1'] = 'Название статьи'
for i in range(len(publication_info)):
    sheet[f'A{i+2}'] = publication_info[i].get("title")
sheet['B1'] = 'Название журнала, номер, год выпуска'
sheet['C1'] = 'Скопус или веб оф сайнс'
sheet['D1'] = 'Рецензирумые зарубежные журналы (РИНЦ и другие через e-library - проверка)'
sheet['E1'] = 'Журналы, рекомендуемые КОКСНВО РК (список журналов)'
sheet['F1'] = 'Название сборников конференции международные/республиканские(название сборника, месяц, год )'
sheet['G1'] = 'Оттиск прикрепить'
sheet['H1'] = 'Количество цитирование Скопус/ веб оф сайнс'
sheet['I1'] = 'Тип автора'
sheet['J1'] = 'Количество авторов'
sheet['D2'] = 'Поиск на сайте или или Название сайте ДА/нет'
sheet['E2'] = 'Поиск на сайте или или Название сайте ДА/нет'
sheet['F2'] = 'Поиск на сайте или или Название сайте ДА/нет'
sheet['I2'] = 'Первый автор Автор кореспондент Соавтор (низбадающий список)'

sheet = sheet1
sheet['A1'] = 'Название Свидетельства'
sheet['B1'] = 'Данные о свидетельстве'
sheet['C1'] = 'авторы'
sheet['D1'] = 'Прикрепить копию пдф'

sheet = sheet2
sheet['A1'] = 'Название '
sheet['B1'] = 'Год выпуска, издательство.'
sheet['C1'] = 'Количество страниц'
sheet['C1'] = 'Прикрепить оттиск пдф'
sheet['A2'] = 'Чтение с вложеноого файла автоматический'
sheet['B2'] = 'Чтение с вложеноого файла автоматический'

# Save the changes to the Excel file
wb.save('data.xlsx')

